#-*- encoding:utf-8 -*-
import os
import sys
import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple
from scipy.ndimage import binary_erosion, binary_dilation, binary_closing, binary_opening
from openpyxl import Workbook
from openpyxl.styles import PatternFill

sys.path.append(str(Path(__file__).parent.parent.parent))

from workflow.compare.griddiff.core.grid_stat_computer import GridStatResult
from workflow.compare.griddiff.plotting.plot_heatmap import plot_statistic_heatmap
from workflow.processor.converter.point_cloud_rasterization import point_cloud_rasterization
from workflow.compare.griddiff.utils.image_processing import *
from utils.logger import logger

plt.rcParams['font.sans-serif'] = ['WenQuanYi Zen Hei']
plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题


class GridComparisonVisualizer:

    def __init__(
        self,
        cfg: object,
        cloud_dict: Dict[str, Dict[str, object]],
        common_las_keys: List[Tuple[float, float]],
        common_keys: List[Tuple[float, float]],
        ground_method: str,
        grid_stat_results: GridStatResult,
        result_subdir: str
    ) -> None:
        """
        参数说明：
        - cloud_dict: {project_name: {'las': PointCloud, 'las_cutter': PointCloudGridCutter, ...}}
        - common_keys: 所有参与对比的公共网格键 (x, y)
        - grid_stat_results: 统计结果对象 (包含 stats 和 diffs)
        - cfg: 配置对象（具有 comp_pair, data_types, plant_thres 等属性
        """
        self.cfg = cfg
        self.result_subdir = result_subdir
        self.cloud_dict = cloud_dict
        self.common_las_keys = common_las_keys
        self.common_keys = common_keys
        self.ground_method = ground_method
        self.stats = grid_stat_results.stats
        self.diffs = grid_stat_results.diffs

        self.lasA_img = None
        self.lasB_img = None
        self.common_lasB_img = None
        self.common_ground_img = None

        self.masks_config = getattr(self.cfg.grid_diff, 'masks', {})
        self.use_morphology = getattr(self.cfg.grid_diff, 'use_morphology', False)

        self.K = len(self.common_keys)

        # self.fit_flags = None

        for project_name, dt_dict in self.stats.items():
            for dt, metric_dict in dt_dict.items():
                for metric_name, v in metric_dict.items():
                    self._has_invalid_values(v, label=f"stats: {project_name} - {dt} - {metric_name}")

        for dt, metric_dict in self.diffs.items():
            for metric_name, v in metric_dict.items():
                self._has_invalid_values(v, label=f"diffs: {dt} - {metric_name}")
        
        # 预处理阶段
        self._prepare_grid_indices()
        self._prepare_images()
        self._prepare_masks()  # 总是执行，final_mask一直可用

        # if self.cfg.grid_diff.spline_fit and grid_stat_results.fit_flags is not None:
        #     self.fit_flags = self.array_to_heatmap(grid_stat_results.fit_flags, flipud=False)

    def _has_invalid_values(self, v, label=""):
        """
        判断数组是否含有 NaN、None 或 inf
        """
        if isinstance(v, cp.ndarray):
            v = cp.asnumpy(v)
        flag = True
        if np.any(np.isnan(v)):
            logger.info(f"⚠️ {label}: 该数组中含有NaN值")
            flag = False

        if np.any(np.isinf(v)):
            logger.info(f"⚠️ {label}: 该数组中含有INF值")
            flag = False

        if np.any([x is None for x in np.ravel(v)]):
            logger.info(f"⚠️ {label}: 该数组中含有None值")
            flag = False
        
        if flag:
            logger.info(f"✅ {label}: 该数组中都是有效值")

    def _prepare_grid_indices(self):
        """计算网格索引与热力图行列映射"""
        x_coords = [xy[0] for xy in self.common_keys]
        y_coords = [xy[1] for xy in self.common_keys]
        self.x_unique = sorted(set(x_coords))
        self.y_unique = sorted(set(y_coords))
        self.nx = len(self.x_unique)
        self.ny = len(self.y_unique)
        self.x2i = {x: i for i, x in enumerate(self.x_unique)}
        self.y2j = {y: j for j, y in enumerate(self.y_unique)}
        # self.y2j = {y: self.ny - 1 - j for j, y in enumerate(self.y_unique)}  # 反转Y轴映射，保证统计数据填充到热力图时原点在左下角，与点云实际XY朝向一致
        self.x_idx = np.array([self.x2i[x] for x, _ in self.common_keys])
        self.y_idx = np.array([self.y2j[y] for _, y in self.common_keys])

    def _prepare_images(self):
        """生成点云图像（LAS A/B + 重叠区域）"""
        self.lasA_img = point_cloud_rasterization(
            pc=self.cloud_dict[self.cfg.comp_pair[0]]['las'],
            sampling_rate=self.cfg.rasterization.sampling_rate,
            sampling_mode=self.cfg.rasterization.sampling_mode,
            enable_fill=False,
            enable_plt=False,
            fill_algorithm='nearest',
            reverse=self.cfg.rasterization.reverse
        )

        self.groundA_img = point_cloud_rasterization(
            pc=self.cloud_dict[self.cfg.comp_pair[0]][self.ground_method],
            sampling_rate=self.cfg.rasterization.sampling_rate,
            sampling_mode=self.cfg.rasterization.sampling_mode,
            enable_fill=False,
            enable_plt=False,
            fill_algorithm='nearest',
            reverse=self.cfg.rasterization.reverse
        )

        self.lasB_img = point_cloud_rasterization(
            pc=self.cloud_dict[self.cfg.comp_pair[1]]['las'],
            sampling_rate=self.cfg.rasterization.sampling_rate,
            sampling_mode=self.cfg.rasterization.sampling_mode,
            enable_fill=False,
            enable_plt=False,
            fill_algorithm='nearest',
            reverse=self.cfg.rasterization.reverse
        )

        self.groundB_img = point_cloud_rasterization(
            pc=self.cloud_dict[self.cfg.comp_pair[1]][self.ground_method],
            sampling_rate=self.cfg.rasterization.sampling_rate,
            sampling_mode=self.cfg.rasterization.sampling_mode,
            enable_fill=False,
            enable_plt=False,
            fill_algorithm='nearest',
            reverse=self.cfg.rasterization.reverse
        )

        self.common_lasB_img = compute_common_regions(self.cfg, 
                                                       self.cloud_dict[self.cfg.comp_pair[1]]['las_cutter'],
                                                       self.common_las_keys)
        self.common_ground_img = compute_common_regions(self.cfg,
                                                        self.cloud_dict[self.cfg.comp_pair[1]][f'{self.cfg.wanted_ground}_cutter'],
                                                        self.common_keys)

    def _prepare_masks(self):
        dt = self.cfg.wanted_ground
        stats = self.stats
        diffs = self.diffs

        self.height_mask = np.zeros((self.ny, self.nx), dtype=bool)
        self.density_mask = np.zeros((self.ny, self.nx), dtype=bool)
        self.vegetation_mask = np.zeros((self.ny, self.nx), dtype=bool)

        self.final_mask = np.ones((self.ny, self.nx), dtype=bool)

        computed_metrics = set(self.cfg.grid_diff.metrics)

        for metric, want_mask in getattr(self, 'masks_config', {}).items():
            if not want_mask:
                continue

            if metric not in computed_metrics:
                logger.warning(f"掩码请求 '{metric}' 但该指标未在 metrics 中计算，跳过")
                continue

            if metric == '高程平均值':
                arr = diffs[dt]['高程平均值']

                non_zero_count = np.count_nonzero(arr != 0)
                total_length = len(arr)
                logger.info(f"高程平均值差异数组：非零值数量 = {non_zero_count} 总长度 = {total_length} ")
                
                # 方式1：所有位置都有效，默认行为
                valid = np.full(arr.shape, True, dtype=bool)
                # 方式2：只有非零差异的才有效（推荐，避免完成相同均值的点云网格参与掩码，这些区域可以不用差分比较）
                # valid = (arr != 0)
                # 方式3：差异绝对值大于或小于某个阈值才有效，能过滤微小/高异常噪声
                # valid = (np.abs(arr) <= 0.5)
                self.height_mask[self.y_idx[valid], self.x_idx[valid]] = True

                # 高程均值的掩码额外做图像形态学处理
                self.height_mask_raw = self.height_mask.copy()
                if self.use_morphology:
                    structure = np.ones((self.cfg.grid_diff.erosion_size, self.cfg.grid_diff.erosion_size), dtype=bool)
                    '''
                    binary_closing: 闭运算，先膨胀再腐蚀，净效果是填补图像中间的小空洞，整体形状基本不变。如果不事先填补空洞，后续腐蚀运算时会削减更多有效区域
                    binary_opening: 开运算，先做腐蚀再做膨胀，净效果是去除图像外部的小噪声块/小孤岛，整体形状基本不变
                    binary_erosion: 仅腐蚀运算，边界整体向内收缩收紧一圈，区域明显变小

                    上述三个步骤需按顺序执行。
                    先做闭运算再做开运算是因为第一步的closing在填洞的同时，可能会把图像周遭的小噪声点也连成片，opening可以再把误连的小噪声块去掉。
                    其实第二步的opening理论上来说是可选的，因为第三步的腐蚀如果算子尺寸足够大，也有可能把图像周遭误连的小噪声块也腐蚀掉，起到第二步的作用
                    实际是否需要第二步的opening以及每一步的算子尺寸是多少，需要专门调试，暂且搁置
                    '''
                    self.height_mask = binary_closing(self.height_mask, structure=structure)
                    self.height_mask = binary_opening(self.height_mask, structure=structure)
                    self.height_mask = binary_erosion(self.height_mask, structure=structure)

                self.final_mask &= self.height_mask

            elif metric == '密度':
                densityA = stats[self.cfg.comp_pair[0]][dt]['密度']
                densityB = stats[self.cfg.comp_pair[1]][dt]['密度']
                valid = (densityA >= self.cfg.grid_diff.density_thres) & \
                        (densityB >= self.cfg.grid_diff.density_thres)
                self.density_mask[self.y_idx[valid], self.x_idx[valid]] = True

                self.final_mask &= self.density_mask

            elif metric == '高程标准差':
                arr = stats[self.cfg.comp_pair[1]][dt]['高程标准差']
                valid = np.abs(arr) < self.cfg.grid_diff.plant_thres
                self.vegetation_mask[self.y_idx[valid], self.x_idx[valid]] = True

                self.final_mask &= self.vegetation_mask

            # 以后新增统计量的掩码类型时，这里继续加 elif 即可

        # self.masked_common_ground_img, _ = apply_mask_to_rgba(self.common_ground_img, np.flipud(self.final_mask))

    def array_to_heatmap(self, arr, flipud, mask=None):
        heatmap = np.zeros((self.ny, self.nx))
        heatmap[self.y_idx, self.x_idx] = arr  # 这个填充顺序非常关键！
        if mask is not None:
            # 支持 bool、uint8(0/1)、float(0.0/1.0)
            heatmap *= mask
        if flipud:  # 图片的reverse=True时，静态热力图需要翻转，plotly不用翻转
            heatmap = np.flipud(heatmap)
        return heatmap
    
    def export_heatmap_to_excel(self, heatmap, filename, threshold=None):
        """
        将热力图矩阵导出到Excel，并根据阈值着色大于阈值的单元格为黄色。

        :param heatmap: 需要导出的热力图矩阵（二维数组）
        :param filename: 输出的Excel文件名
        :param threshold: 可选的阈值，默认使用最大值的80%作为阈值
        """
        # 计算阈值，默认使用最大值的80%
        if threshold is None:
            threshold = np.max(heatmap) * 0.8

        wb = Workbook()
        ws = wb.active
        ws.title = "热力图矩阵"

        # 填充 Excel 表格，按行按列顺序填充矩阵数据
        for i in range(heatmap.shape[0]):
            for j in range(heatmap.shape[1]):
                # 填充单元格
                value = heatmap[i, j]
                cell = ws.cell(row=i+1, column=j+1, value=value)

                # 如果值大于阈值，设置单元格背景色为黄色
                if value > threshold:
                    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.fill = yellow_fill

        wb.save(filename)
        logger.info(f"热力图已导出到 {filename}")

    def _plot_masks(self):
        """绘制掩码图集：第一张固定 4 张点云鸟瞰图，第二张动态绘制所有存在的掩码"""

        # 第一张图：固定 4 张点云鸟瞰图
        fig1, axes1 = plt.subplots(2, 2, figsize=(16, 16))
        plot_statistic_heatmap(self.lasA_img, fig1, axes1.flat[0], '点云A', with_colorbar=False, origin='upper')
        plot_statistic_heatmap(self.lasB_img, fig1, axes1.flat[1], '点云B', with_colorbar=False, origin='upper')
        plot_statistic_heatmap(self.groundA_img, fig1, axes1.flat[2], '地面点A', with_colorbar=False, origin='upper')
        plot_statistic_heatmap(self.groundB_img, fig1, axes1.flat[3], '地面点B', with_colorbar=False, origin='upper')
        if self.cfg.save:
            plt.savefig(os.path.join(self.result_subdir, "点云鸟瞰图对比.png"), dpi=300, bbox_inches='tight', pad_inches=0.2)
        plt.tight_layout()
        if self.cfg.grid_diff.show_fig: plt.show()
        plt.close(fig1)

        # 第二张图：动态绘制所有存在的掩码
        mask_plots = []  # 收集所有已生成的掩码及其标题

        # 高程差异掩码（原始和处理后）
        if hasattr(self, 'height_mask_raw') and np.any(self.height_mask_raw):
            mask_plots.append((self.height_mask_raw, '高程差异蒙版 (原始)'))
        if hasattr(self, 'height_mask') and np.any(self.height_mask):
            mask_plots.append((self.height_mask, '高程差异蒙版 (预处理后)'))

        # 密度掩码
        if hasattr(self, 'density_mask') and np.any(self.density_mask):
            title = f"密度蒙版 (网格点数量 > {self.cfg.grid_diff.density_thres})"
            mask_plots.append((self.density_mask, title))

        # 植被掩码
        if hasattr(self, 'vegetation_mask') and np.any(self.vegetation_mask):
            title = f"浅植蒙版 (点云B的标准差 < {self.cfg.grid_diff.plant_thres} 米)"
            mask_plots.append((self.vegetation_mask, title))

        # 最终掩码
        if hasattr(self, 'final_mask') and np.any(self.final_mask < 1):  # 如果不是全True才画
            mask_plots.append((self.final_mask, '最终蒙版'))

        # 如果没有任何掩码可画，就跳过
        if not mask_plots:
            logger.warning("没有生成任何掩码，跳过掩码汇总图")
            return

        # 动态布局：尽量接近方形，列数最多4
        n_masks = len(mask_plots)
        cols = min(4, n_masks)
        rows = (n_masks + cols - 1) // cols
        fig2, axes2 = plt.subplots(rows, cols, figsize=(4*cols, 4*rows + 2))
        axes2 = axes2.flat if n_masks > 1 else [axes2]  # 兼容单图情况

        for i, (mask_img, title) in enumerate(mask_plots):
            plot_statistic_heatmap(mask_img, fig2, axes2[i], title, with_colorbar=False, origin='lower', cmap='binary_r')

        # 隐藏可能多余的子图
        for j in range(i+1, len(axes2)):
            axes2[j].axis('off')

        if self.cfg.save:
            plt.savefig(os.path.join(self.result_subdir, "蒙版汇总.png"), dpi=300, bbox_inches='tight', pad_inches=0.2)
        plt.tight_layout()
        if self.cfg.grid_diff.show_fig:
            plt.show()
        plt.close(fig2)

    def _plot_diff_hms(self, metric, dt):
        """绘制单一统计量的差分热力图"""
        diff_arr = self.diffs[dt][metric]
        heatmapD = self.array_to_heatmap(diff_arr, flipud=False, mask=self.final_mask)

        fig, axes = plt.subplots(1, 2, figsize=(16, 8))  # 宽，高
        plot_statistic_heatmap(heatmapD, fig, axes[0], f"{metric}差异({dt.lower()})", with_colorbar=True, origin='lower', underlay_rgb=None)
        plot_statistic_heatmap(heatmapD, fig, axes[1], f"{metric}差异叠加点云图像({dt.lower()})", with_colorbar=True, origin='lower', underlay_rgb=self.common_lasB_img)

        if self.cfg.save:
            save_path = os.path.join(self.result_subdir, f"{metric}_{dt}.png")
            plt.savefig(save_path, dpi=300, bbox_inches='tight', pad_inches=0.2)

        plt.tight_layout()
        if self.cfg.grid_diff.show_fig: plt.show()
        plt.close(fig)

    def visualize(self):
        """可视化入口"""
        self._plot_masks()
        for metric in self.cfg.grid_diff.metrics:
            if metric in self.stats[self.cfg.comp_pair[0]][self.cfg.wanted_ground]:
                self._plot_diff_hms(metric, self.cfg.wanted_ground)
        plt.close('all')  # 兜底释放内存资源，防止内存泄漏
        
        self.raw_height_heatmap = self.array_to_heatmap(self.diffs[self.cfg.wanted_ground]['高程平均值'], flipud=False, mask=self.final_mask)
    
        if self.cfg.save:
            self.export_heatmap_to_excel(self.raw_height_heatmap, os.path.join(self.result_subdir, '导出高程差异.xlsx'))